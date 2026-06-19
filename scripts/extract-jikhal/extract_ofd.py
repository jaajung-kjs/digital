"""OFD 선번장 → cores/opgw 원시 JSON.

직할_OFD선번장_260203.xlsx 의 13 국소 시트만 순회. 블록(상대국/T/L)별 코어 대장.
산출: out/ofd_cores.json (사용 코어), out/ofd_opgw.json (직할-직할 OPGW 쌍, dedup).
"""
import json
import os
import openpyxl
from normalize import normalize_station, is_direct, pair_key

SRC = "/Users/jsk/1210/digital/직할/직할_OFD선번장_260203.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "out")

# 시트명 → 직할 subKey ((구)/(신)춘천 구분 보존 — 셀용 normalize_station 과 별개)
SHEET_SUB = {
    "(구)춘천": "guchuncheon", "(신)춘천": "sinchuncheon", "북춘천": "bukchuncheon",
    "남춘천": "namchuncheon", "서홍천": "seohongcheon", "열병합": "yeolbyeonghap",
    "인제": "inje", "양구": "yanggu", "철원": "cheorwon", "화천": "hwacheon",
    "화천HP": "hwacheonhp", "춘천HP": "chuncheonhp", "소양HP": "soyanghp",
}


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def splice_of(text):
    t = text or ""
    for k in ("단선", "융착", "패치"):
        if k in t:
            return k
    return None


def main():
    os.makedirs(OUT, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    cores = []
    blocks = []  # [{subKey, block, peerKey, peerRaw, cores}] — 슬롯(케이블) 단위
    for sheet, subKey in SHEET_SUB.items():
        if sheet not in wb.sheetnames:
            print(f"  ⚠️ 시트 없음: {sheet}")
            continue
        ws = wb[sheet]
        in_data = False
        peer_raw = peer_key = None
        block_cores = 0  # 현재 블록 총 코어수(OPGW 용량)
        block_idx = 0    # 시트 내 블록(케이블) 순번 — 슬롯 단위
        for row in ws.iter_rows(values_only=True):
            a = row[0] if len(row) > 0 else None
            # 블록 경계: 섹션 헤더 / 메타. 블록(슬롯) 단위는 '상대국(B) 바뀔 때' (아래).
            if isinstance(a, str) and a.strip() == "No.":
                in_data = True
                continue
            if isinstance(a, str) and ("관리사업소" in a or "장소" in a):
                in_data = False
                if peer_key and block_cores:  # 직전 블록 기록
                    blocks.append({"subKey": subKey, "block": block_idx, "peerKey": peer_key,
                                   "peerRaw": peer_raw, "cores": block_cores})
                peer_raw = peer_key = None
                block_cores = 0
                continue
            if not in_data:
                continue
            core = None
            try:
                core = int(a)
            except (TypeError, ValueError):
                continue  # 데이터 행 아님
            # B(상대국명) 가 새로 나오면 = 새 케이블(슬롯) 시작 → 직전 블록 flush 후 block_idx++
            b = row[1] if len(row) > 1 else None
            if b and str(b).strip():
                if peer_key and block_cores:
                    blocks.append({"subKey": subKey, "block": block_idx, "peerKey": peer_key,
                                   "peerRaw": peer_raw, "cores": block_cores})
                block_idx += 1
                block_cores = 0
                peer_raw = str(b).strip()
                peer_key = normalize_station(peer_raw)
            block_cores += 1
            purpose = (str(row[2]).strip() if len(row) > 2 and row[2] else None)
            circuit = (str(row[3]).strip() if len(row) > 3 and row[3] else None)
            usage_raw = row[13] if len(row) > 13 else None
            # 사용여부: N==1→사용, purpose/circuit 있는데 blank→사용, N==0&무purpose→예비(None)
            if num(usage_raw) == 1:
                usage = "사용"
            elif (purpose or circuit) and usage_raw in (None, ""):
                usage = "사용"
            elif num(usage_raw) == 0:
                usage = None
            else:
                usage = None
            # 사용 코어만 기록(예비 제외)
            if usage != "사용" and not purpose:
                continue
            cores.append({
                "subKey": subKey, "block": block_idx, "core": core, "peerRaw": peer_raw, "peerKey": peer_key,
                "purpose": purpose, "circuitText": circuit,
                "spliceType": splice_of(circuit), "usageOverride": usage,
                "loss1310": num(row[5]) if len(row) > 5 else None,
                "dist1310": num(row[6]) if len(row) > 6 else None,
                "loss1550": num(row[7]) if len(row) > 7 else None,
                "dist1550": num(row[8]) if len(row) > 8 else None,
                "inspectResult": (str(row[9]).strip() if len(row) > 9 and row[9] else None),
            })
        # 시트 끝 블록 flush
        if peer_key and block_cores:
            blocks.append({"subKey": subKey, "block": block_idx, "peerKey": peer_key,
                           "peerRaw": peer_raw, "cores": block_cores})

    json.dump(cores, open(os.path.join(OUT, "ofd_cores.json"), "w"), ensure_ascii=False, indent=1)
    json.dump(blocks, open(os.path.join(OUT, "ofd_blocks.json"), "w"), ensure_ascii=False, indent=1)
    direct_blocks = sum(1 for b in blocks if is_direct(b["peerKey"]))
    print(f"cores {len(cores)}  blocks {len(blocks)} (직할대국 블록 {direct_blocks})")


if __name__ == "__main__":
    main()
